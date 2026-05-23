import streamlit as st
import pandas as pd
import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components

# Настройки страницы
st.set_page_config(page_title="Архитектор расписания v2.0 (ИДЕАЛ)", layout="wide", page_icon="📅", initial_sidebar_state="collapsed")

# --- CSS МАГИЯ ДЛЯ ЧИСТОГО ИНТЕРФЕЙСА ---
st.markdown("""
<style>
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    .stAppDeployButton {display: none;}
    [data-testid="stSidebar"] { display: none !important; }
    [data-testid="collapsedControl"] { display: none !important; }
    
    button[data-baseweb="tab"] {
        font-size: 16px !important;
        font-weight: bold !important;
        padding: 12px 24px !important;
    }
    
    .custom-footer {
        position: fixed; left: 0; bottom: 0; width: 100%;
        background-color: #ffffff; color: #888; text-align: right;
        padding: 10px 20px; font-size: 14px; font-weight: bold; z-index: 100;
        border-top: 1px solid #eee;
    }
</style>
<div class="custom-footer">сделано Lottarruu | v2.0 Dynamic Engine</div>
""", unsafe_allow_html=True)

# --- ИНИЦИАЛИЗАЦИЯ БАЗЫ ДАННЫХ ---
if 'db_rooms' not in st.session_state: st.session_state.db_rooms = []
if 'db_teachers' not in st.session_state: st.session_state.db_teachers = []
if 'global_groups' not in st.session_state: st.session_state.global_groups = []
if 'custom_holidays' not in st.session_state: st.session_state.custom_holidays = []

if 'editing_idx' not in st.session_state: st.session_state.editing_idx = None
if 'draft_group' not in st.session_state: st.session_state.draft_group = None
if 'temp_subjects' not in st.session_state: st.session_state.temp_subjects = []
if 'do_scroll' not in st.session_state: st.session_state.do_scroll = False
if 'sub_error' not in st.session_state: st.session_state.sub_error = False

# --- CALLBACK ДЛЯ БЕЗОПАСНОГО ДОБАВЛЕНИЯ ПРЕДМЕТА ---
def add_subject_cb():
    name = st.session_state.sched_sub_name.strip()
    t_name = st.session_state.sched_sub_teacher
    hours = st.session_state.sched_sub_hours
    accent = st.session_state.sched_sub_accent
    
    invalid_teachers = ["Сначала добавьте учителей!", "Нет учителей для этого предмета"]
    
    if name and t_name not in invalid_teachers:
        st.session_state.temp_subjects.append({'name': name, 'teacher': t_name, 'hours': hours, 'accent': accent})
        st.session_state.sched_sub_name = ""
        st.session_state.sched_sub_accent = False
        st.session_state.sub_error = False
    else:
        st.session_state.sub_error = True

# --- ЛОГИКА ПРАЗДНИКОВ (СКОРРЕКТИРОВАННАЯ) ---
def _is_holiday(date_obj):
    # Жесткий перехват: эти дни всегда рабочие в нашем расписании!
    if (date_obj.month == 11 and date_obj.day in [7, 8]) or (date_obj.month == 2 and date_obj.day == 23):
        return None 

    if date_obj in st.session_state.custom_holidays: return "Выходной по решению завуча"
    if date_obj.month == 1 and 1 <= date_obj.day <= 12: return "Зимние каникулы"
    if date_obj.month == 5 and 1 <= date_obj.day <= 9: return "Майские каникулы"
    if date_obj.month == 3 and date_obj.day == 21: return "Нооруз"
    
    floating_holidays = {
        2025: [(3, 30, "Орозо айт"), (6, 6, "Курман айт")], 2026: [(3, 20, "Орозо айт"), (5, 27, "Курман айт")],
        2027: [(3, 10, "Орозо айт"), (5, 17, "Курман айт")], 2028: [(2, 27, "Орозо айт"), (5, 5, "Курман айт")],
        2029: [(2, 15, "Орозо айт"), (4, 24, "Курман айт")], 2030: [(2, 4, "Орозо айт"), (4, 14, "Курман айт")]
    }
    for m, d, name in floating_holidays.get(date_obj.year, []):
        if date_obj.month == m and date_obj.day == d: return name
    return None

# --- НОВЫЙ ДИНАМИЧЕСКИЙ ДВИЖОК ГЕНЕРАЦИИ (V2.0) ---
def generate_master_schedule():
    all_dates = set()
    flat_groups = []
    results, balances, group_max_cols = {}, {}, {}
    
    # 1. Распаковка групп и расчет рабочих дней
    for g in st.session_state.global_groups:
        for c in range(1, g['num_courses'] + 1):
            c_name = f"{g['name']} ({c} курс)"
            
            start_year = g['start_date'].year + (c - 1)
            end_year = g['start_date'].year + c
            c_start_date = datetime.date(start_year, 9, 1)
            c_end_date = datetime.date(end_year, 5, 25) # Если часов слишком много, движок сам растянет дату!
            
            # Считаем точные рабочие дни курса
            group_working_days = []
            curr = c_start_date
            # Берем с запасом до 15 июня, чтобы движку было куда растягивать расписание, если часов слишком много
            while curr <= datetime.date(end_year, 6, 15): 
                all_dates.add(curr)
                holiday = _is_holiday(curr)
                wd = curr.weekday()
                is_work_day = True
                if holiday or wd == 6 or (wd == 5 and g['work_week'] == "5-дневная рабочая неделя") or curr.month in [7,8]:
                    is_work_day = False
                if is_work_day:
                    group_working_days.append(curr)
                curr += datetime.timedelta(days=1)
                
            # Резерв завуча — жестко последние 4 рабочих дня в мае!
            may_working_days = [d for d in group_working_days if d <= c_end_date]
            reserve_days = set(may_working_days[-4:]) if len(may_working_days) > 4 else set()
            
            c_bal = {}
            for sub in g['subjects']:
                base = sub['hours'] // g['num_courses']
                rem = sub['hours'] % g['num_courses']
                actual_hours = base + (rem if c == g['num_courses'] else 0)
                
                c_bal[sub['name']] = {
                    'h': actual_hours, 't': sub['teacher'], 
                    'accent': sub['accent'], 'daily_usage': 0
                }
            
            flat_groups.append({
                'id': c_name, 'ref': g, 'bal': c_bal, 'cal': {}, 
                'max_w': g['max_weekday_lessons'], 'max_s': g['max_saturday_lessons'], 
                'max_per_sub': g['max_per_subject'], 'g_room': g['room'],
                'reserve_days': reserve_days,
                'c_start': c_start_date, 'c_end': datetime.date(end_year, 6, 15),
                'rem_days': max(1, len(group_working_days) - len(reserve_days)) # Для расчета срочности
            })
            balances[c_name] = {}
            group_max_cols[c_name] = max(g['max_weekday_lessons'], g['max_saturday_lessons'])

    all_dates = sorted(list(all_dates))

    # 2. Глобальный распределительный цикл (Движок Жадины)
    for date_obj in all_dates:
        is_spring = (1 <= date_obj.month <= 5)
        holiday = _is_holiday(date_obj)
        weekday = date_obj.weekday()
        
        daily_teacher_usage, daily_room_usage = {}, {}
        active_today = []
        
        for fg in flat_groups:
            if fg['c_start'] <= date_obj <= fg['c_end']:
                # Если все часы у группы УЖЕ кончились, и дата перевалила за 25 мая - просто стопаем для неё
                total_h_left = sum(s['h'] for s in fg['bal'].values())
                if total_h_left == 0 and date_obj > datetime.date(date_obj.year, 5, 25):
                    continue

                if holiday: 
                    fg['cal'][date_obj] = {"status": "Праздник", "info": holiday, "slots": []}
                elif weekday == 6 or (weekday == 5 and fg['ref']['work_week'] == "5-дневная рабочая неделя") or date_obj.month in [7,8]: 
                    pass 
                elif date_obj in fg['reserve_days']:
                    max_slots = fg['max_s'] if weekday == 5 else fg['max_w']
                    fg['cal'][date_obj] = {
                        "status": "Резерв завуча", 
                        "slots": [{"sub": "Резерв", "conflict": False} for _ in range(max_slots)]
                    }
                else:
                    fg['rem_days'] -= 1 # Один день прошел
                    max_slots = fg['max_s'] if weekday == 5 else fg['max_w']
                    for s_name in fg['bal']: fg['bal'][s_name]['daily_usage'] = 0
                        
                    active_today.append({
                        'fg': fg, 'max_slots': max_slots, 'schedule': [None] * max_slots
                    })
                    fg['cal'][date_obj] = {"status": "Рабочий", "slots": []}
        
        if not active_today: continue
            
        max_slots_today = max([ag['max_slots'] for ag in active_today])
        
        # Распределяем уроки в дне
        for slot_idx in range(max_slots_today):
            if slot_idx not in daily_teacher_usage: daily_teacher_usage[slot_idx] = set()
            if slot_idx not in daily_room_usage: daily_room_usage[slot_idx] = set()
            
            for ag in active_today:
                if slot_idx >= ag['max_slots']: continue
                if ag['schedule'][slot_idx] is not None: continue 
                
                fg = ag['fg']
                is_pair_start = (slot_idx % 2 == 0) and (slot_idx < 6) and (slot_idx + 1 < ag['max_slots'])
                
                # Подсчет выживших предметов
                active_subs_count = sum(1 for s in fg['bal'].values() if s['h'] > 0)
                # Если предметов мало или год заканчивается, а часов гора — СНИМАЕМ ЛИМИТЫ (Анти-облысение)
                emergency_mode = (active_subs_count <= 3) or (fg['rem_days'] < 30 and sum(s['h'] for s in fg['bal'].values()) > fg['rem_days'] * 5)
                
                valid_subs = []
                for s_name, s_data in fg['bal'].items():
                    if s_data['h'] <= 0: continue
                    t_name = s_data['t']
                    if t_name in daily_teacher_usage[slot_idx]: continue
                    
                    # Проверяем лимиты
                    if not emergency_mode and s_data['daily_usage'] >= fg['max_per_sub']: continue
                    
                    # Может ли стать парой?
                    can_be_pair = False
                    if is_pair_start and s_data['h'] >= 2:
                        if not emergency_mode and s_data['daily_usage'] + 2 > fg['max_per_sub']:
                            pass # пара нарушит лимит
                        else:
                            future_t_usage = daily_teacher_usage.get(slot_idx+1, set())
                            if t_name not in future_t_usage: can_be_pair = True
                            
                    valid_subs.append((s_name, can_be_pair))
                
                chosen_sub = None
                will_be_pair = False
                
                if valid_subs:
                    def get_urgency(item):
                        s_name, can_pair = item
                        s_data = fg['bal'][s_name]
                        # Коэффициент срочности = остаток часов / оставшиеся дни (идеальный балансировщик)
                        urgency = s_data['h'] / max(1, fg['rem_days'])
                        if is_spring and s_data['accent']: urgency += 2.0
                        if is_pair_start and can_pair: urgency += 100.0 # Пары важнее всего
                        # Штрафуем, если предмет уже был сегодня, чтобы раскидывать их по дню
                        urgency -= (s_data['daily_usage'] * 5) 
                        return urgency
                        
                    valid_subs.sort(key=get_urgency, reverse=True)
                    chosen_sub, will_be_pair = valid_subs[0]
                
                if chosen_sub:
                    t_name = fg['bal'][chosen_sub]['t']
                    t_obj = next((t for t in st.session_state.db_teachers if t['name'] == t_name), None)
                    target_room = t_obj['room'] if t_obj and t_obj['room'] != "Нет" else fg['g_room']
                    
                    room_conflict = False
                    if target_room != "Нет":
                        if target_room in daily_room_usage[slot_idx]: room_conflict = True
                        else: daily_room_usage[slot_idx].add(target_room)
                    
                    # Ставим первый урок
                    ag['schedule'][slot_idx] = {"sub": chosen_sub, "conflict": room_conflict}
                    fg['bal'][chosen_sub]['h'] -= 1
                    fg['bal'][chosen_sub]['daily_usage'] += 1
                    daily_teacher_usage[slot_idx].add(t_name)
                    
                    # Ставим второй урок (пару)
                    if is_pair_start and will_be_pair:
                        if slot_idx + 1 not in daily_teacher_usage: daily_teacher_usage[slot_idx + 1] = set()
                        if slot_idx + 1 not in daily_room_usage: daily_room_usage[slot_idx + 1] = set()
                        
                        ag['schedule'][slot_idx+1] = {"sub": chosen_sub, "conflict": room_conflict}
                        fg['bal'][chosen_sub]['h'] -= 1
                        fg['bal'][chosen_sub]['daily_usage'] += 1
                        daily_teacher_usage[slot_idx+1].add(t_name)
                        if target_room != "Нет": daily_room_usage[slot_idx+1].add(target_room)
                else:
                    # Если никто не подошел, значит реально дыра
                    ag['schedule'][slot_idx] = {"sub": "—", "conflict": False}

        for ag in active_today: 
            ag['fg']['cal'][date_obj]['slots'] = ag['schedule']

    for fg in flat_groups: 
        balances[fg['id']] = {k: v['h'] for k, v in fg['bal'].items()}
        results[fg['id']] = (fg['cal'], balances[fg['id']], group_max_cols[fg['id']])
        
    return results

# --- ЭКСПОРТ В EXCEL ---
def create_excel(all_results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    font_t = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_d = Font(name="Calibri", size=11)
    font_bd = Font(name="Calibri", size=11, bold=True)
    f_h = PatternFill(start_color="1F4E78", fill_type="solid")
    f_hol = PatternFill(start_color="FCE4D6", fill_type="solid")
    f_wk = PatternFill(start_color="F2F2F2", fill_type="solid")
    f_err = PatternFill(start_color="FF0000", fill_type="solid")
    border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
    al_c = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_r = wb.create_sheet("Остаток часов")
    ws_r.append(["Группа", "Предмет", "Остаток (ч)"])
    for r_idx in range(1, 4): ws_r.cell(row=1, column=r_idx).font = font_h; ws_r.cell(row=1, column=r_idx).fill = f_h
    
    for t_name, (cal, rem, max_c) in all_results.items():
        for s_name, h in rem.items():
            if h > 0: ws_r.append([t_name, s_name, h])
            
        ws = wb.create_sheet(t_name[:31])
        m_c = max(1, max_c)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2+m_c)
        ws["A1"].value, ws["A1"].font, ws["A1"].fill, ws["A1"].alignment = f"Расписание: {t_name}", font_t, f_h, al_c
        
        headers = ["Дата", "День недели"] + [f"Урок {i+1}" for i in range(m_c)]
        ws.append(headers)
        for i, _ in enumerate(headers, 1): 
            ws.cell(row=2, column=i).font = font_h
            ws.cell(row=2, column=i).fill = f_h
            ws.cell(row=2, column=i).alignment = al_c
            ws.cell(row=2, column=i).border = border
        
        r_idx = 3
        # Чтобы не выводить длинные пустые хвосты в июне, если часы кончились раньше:
        last_active_date = max([d for d, info in cal.items() if any(s["sub"] != "—" for s in info["slots"])], default=list(cal.keys())[0])
        
        for d in sorted(cal.keys()):
            if d > last_active_date and d > datetime.date(d.year, 5, 25): 
                break # Обрезаем вывод пустоты после 25 мая, если всё вычитано
            
            info = cal[d]
            ws.cell(row=r_idx, column=1, value=d.strftime("%d.%m.%Y")).alignment = al_c
            ws.cell(row=r_idx, column=2, value=["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"][d.weekday()]).alignment = al_c
            ws.cell(row=r_idx, column=1).font = ws.cell(row=r_idx, column=2).font = font_bd
            
            if info["status"] == "Праздник" or info["status"] == "Резерв завуча":
                ws.merge_cells(start_row=r_idx, start_column=3, end_row=r_idx, end_column=2+m_c)
                icon = "🎉" if info["status"] == "Праздник" else "🔒"
                text = info.get("info", "Резерв завуча")
                c = ws.cell(row=r_idx, column=3, value=f"{icon} {text}")
                c.alignment, c.font = al_c, font_bd
                for col in range(1, 3+m_c): 
                    ws.cell(row=r_idx, column=col).fill = f_hol
                    ws.cell(row=r_idx, column=col).border = border
            else:
                for s_idx in range(m_c):
                    col = 3 + s_idx
                    if s_idx < len(info["slots"]):
                        slot = info["slots"][s_idx]
                        if slot is None: slot = {"sub": "—", "conflict": False}
                        c = ws.cell(row=r_idx, column=col, value=slot["sub"])
                        if slot["conflict"]: c.fill = f_err 
                        elif d.weekday() == 5: c.fill = f_wk
                    else:
                        c = ws.cell(row=r_idx, column=col, value="—")
                        c.alignment = al_c
                        if d.weekday() == 5: c.fill = f_wk
                    c.border = border
            r_idx += 1
            
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max(len(str(c.value or '')) for c in col) + 3, 12)

    for col in ws_r.columns: ws_r.column_dimensions[get_column_letter(col[0].column)].width = 25
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# --- НАВИГАЦИЯ ---
st.markdown("<h2 style='text-align: center; color: #1F4E78;'>📅 Архитектор Расписания ПЛ №14</h2>", unsafe_allow_html=True)
st.write("") 

tab_groups, tab_teachers, tab_rooms, tab_holidays = st.tabs([
    "🏢 Расписание и Группы", 
    "👨‍🏫 База Учителей", 
    "🚪 База Кабинетов", 
    "⚙️ Выходные дни"
])

# =========================================================
# ВКЛАДКА: БАЗА КАБИНЕТОВ
# =========================================================
with tab_rooms:
    st.title("🚪 Управление кабинетами")
    with st.form("add_room", clear_on_submit=True):
        r_name = st.text_input("Название/Номер кабинета")
        if st.form_submit_button("Добавить кабинет") and r_name:
            if r_name not in st.session_state.db_rooms: 
                st.session_state.db_rooms.append(r_name)
                st.success(f"Кабинет '{r_name}' добавлен!")
                st.rerun()
    
    if st.session_state.db_rooms:
        st.markdown("### 📋 Список сохраненных кабинетов:")
        for i, r in enumerate(st.session_state.db_rooms):
            c1, c2 = st.columns([4, 1])
            c1.write(f"Кабинет: **{r}**")
            if c2.button("Удалить", key=f"del_r_{i}"):
                st.session_state.db_rooms.pop(i)
                st.rerun()
    else:
        st.info("Кабинетов пока нет в базе. Заполните форму выше.")

# =========================================================
# ВКЛАДКА: БАЗА УЧИТЕЛЕЙ
# =========================================================
with tab_teachers:
    st.title("👨‍🏫 Управление преподавателями")
    with st.form("add_teacher", clear_on_submit=True):
        t_name = st.text_input("ФИО Преподавателя")
        t_subs = st.text_input("Предметы (через запятую)", placeholder="Алгебра, Геометрия")
        t_room = st.selectbox("Прикрепленный кабинет (Приоритет)", ["Нет"] + st.session_state.db_rooms)
        if st.form_submit_button("Сохранить преподавателя") and t_name:
            st.session_state.db_teachers.append({"name": t_name, "subs": [s.strip() for s in t_subs.split(",")], "room": t_room})
            st.success(f"Учитель {t_name} успешно добавлен!")
            st.rerun()
            
    if st.session_state.db_teachers:
        st.markdown("### 📋 Список зарегистрированных преподавателей:")
        for i, t in enumerate(st.session_state.db_teachers):
            c1, c2, c3 = st.columns([3, 3, 1])
            c1.write(f"**{t['name']}** (Каб: {t['room']})")
            c2.write(f"Предметы: {', '.join(t['subs'])}")
            if c3.button("Удалить", key=f"del_t_{i}"):
                st.session_state.db_teachers.pop(i)
                st.rerun()
    else:
        st.info("Преподаватели отсутствуют.")

# =========================================================
# ВКЛАДКА: ВЫХОДНЫЕ ДНИ
# =========================================================
with tab_holidays:
    st.title("⚙️ Пользовательские нерабочие дни")
    st.write("Добавьте даты, которые будут отмечены как выходные для ВСЕХ групп.")
    
    with st.form("add_holiday_form", clear_on_submit=True):
        new_date = st.date_input("Выберите дату для выходного", value=datetime.date.today())
        if st.form_submit_button("➕ Добавить выходной"):
            if new_date not in st.session_state.custom_holidays:
                st.session_state.custom_holidays.append(new_date)
                st.success(f"Дата {new_date.strftime('%d.%m.%Y')} успешно добавлена!")
                st.rerun()
            else:
                st.warning("Эта дата уже есть в списке выходных.")
                
    if st.session_state.custom_holidays:
        st.markdown("### 🗓️ Список глобальных выходных:")
        st.session_state.custom_holidays.sort()
        for i, d in enumerate(st.session_state.custom_holidays):
            c1, c2 = st.columns([4, 1])
            c1.markdown(f"🛑 **{d.strftime('%d.%m.%Y')}** — Выходной день")
            if c2.button("Удалить", key=f"del_hol_{i}"):
                st.session_state.custom_holidays.pop(i)
                st.rerun()
    else:
        st.info("Пользовательских праздничных дней пока не добавлено.")

# =========================================================
# ВКЛАДКА: РАСПИСАНИЕ И ГРУППЫ
# =========================================================
with tab_groups:
    with st.expander("📊 Управление базой и Генерация", expanded=True):
        if st.session_state.global_groups:
            st.success(f"В базе сохранено групп: {len(st.session_state.global_groups)}")
            for idx, g in enumerate(st.session_state.global_groups):
                c1, c2, c3, c4 = st.columns([3, 2, 1, 1])
                c1.markdown(f"**{g['name']}** ({g['num_courses']} курс)")
                c2.markdown(f"Нагрузка: {sum([s['hours'] for s in g['subjects']])} ч.")
                if c3.button("✏️ Изменить", key=f"ed_g_{idx}"):
                    if st.session_state.editing_idx is None and st.session_state.temp_subjects:
                        st.session_state.draft_group = list(st.session_state.temp_subjects)
                    st.session_state.editing_idx = idx
                    st.session_state.temp_subjects = g['subjects'].copy()
                    st.session_state.do_scroll = True
                    st.rerun()
                if c4.button("❌ Удалить", key=f"del_g_{idx}"):
                    st.session_state.global_groups.pop(idx); st.rerun()
            
            st.markdown("---")
            if st.button("🚀 СГЕНЕРИРОВАТЬ EXCEL-ФАЙЛ", type="primary", use_container_width=True):
                if not st.session_state.db_teachers: 
                    st.error("⚠️ База учителей пуста! Добавьте учителей.")
                else:
                    with st.spinner("Синхронная балансировка расписания (V2.0)..."):
                        excel_data = create_excel(generate_master_schedule())
                        st.download_button("📥 СКАЧАТЬ ГОТОВОЕ РАСПИСАНИЕ", data=excel_data, file_name="Raspisanie_FINAL_V2.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    st.markdown('<div id="edit-section" style="padding-top: 30px;"></div>', unsafe_allow_html=True)
    is_ed = st.session_state.editing_idx is not None
    st.subheader("✏️ " + ("Редактирование группы" if is_ed else "Создание новой группы"))
    
    if st.session_state.draft_group and not is_ed:
        if st.button("Восстановить несохраненный черновик 📝"):
            st.session_state.temp_subjects = list(st.session_state.draft_group)
            st.session_state.draft_group = None; st.rerun()

    cur_g = st.session_state.global_groups[st.session_state.editing_idx] if is_ed else {}
    
    cl, cr = st.columns([1, 1], gap="large")
    with cl:
        st.markdown("##### Параметры группы")
        g_name = st.text_input("Название", value=cur_g.get('name', ''))
        g_c = st.number_input("Курсы", 1, 4, cur_g.get('num_courses', 1))
        
        default_room_idx = 0
        if cur_g.get('room', 'Нет') in ["Нет"] + st.session_state.db_rooms:
            default_room_idx = (["Нет"] + st.session_state.db_rooms).index(cur_g.get('room', 'Нет'))
        g_room = st.selectbox("Прикрепленный кабинет группы", ["Нет"] + st.session_state.db_rooms, index=default_room_idx)
        
        g_week = st.selectbox("Режим", ["5-дневная рабочая неделя", "6-дневная рабочая неделя"], index=0 if cur_g.get('work_week', '') == "5-дневная рабочая неделя" else 1)
        g_w_l = st.slider("Будни (макс. уроков в день)", 4, 8, cur_g.get('max_weekday_lessons', 7)) # 7 уроков в норме
        g_s_l = st.slider("Суббота (макс. уроков)", 1, 8, cur_g.get('max_saturday_lessons', 4)) if g_week == "6-дневная рабочая неделя" else 0
        g_m_s = st.slider("Макс. 1 предмета в день", 1, 8, cur_g.get('max_per_subject', 2))
        
        default_dates = (cur_g.get('start_date', datetime.date(2025,9,1)), cur_g.get('end_date', datetime.date(2026,5,25)))
        dates = st.date_input("Период обучения", value=default_dates)
        
    with cr:
        st.markdown("##### Предметы и Нагрузка")
        with st.container(border=True):
            if st.session_state.sub_error:
                st.error("⚠️ Укажите предмет и выберите подходящего учителя!")
                
            s_name = st.text_input("Название дисциплины (введите и кликните вне поля)", key="sched_sub_name")
            
            # ФИЛЬТРАЦИЯ УЧИТЕЛЕЙ
            if not st.session_state.db_teachers:
                teacher_options = ["Сначала добавьте учителей!"]
            else:
                if s_name:
                    valid_t = [
                        t['name'] for t in st.session_state.db_teachers 
                        if s_name.lower().strip() in [sub.lower().strip() for sub in t.get('subs', [])]
                    ]
                    if valid_t: teacher_options = valid_t
                    else: teacher_options = ["Нет учителей для этого предмета"]
                else:
                    teacher_options = [t['name'] for t in st.session_state.db_teachers]
            
            s_t = st.selectbox("Преподаватель", teacher_options, key="sched_sub_teacher")
            s_h = st.number_input("Общее кол-во часов (1 кредит = 30 ч)", min_value=1, key="sched_sub_hours")
            s_a = st.checkbox("Акцент (на весну)", key="sched_sub_accent")
            
            st.button("➕ Добавить предмет в список", on_click=add_subject_cb, use_container_width=True, key="sched_add_subject_final_btn")

        if st.session_state.temp_subjects:
            st.markdown("**Список дисциплин группы:**")
            
        for i, sub in enumerate(st.session_state.temp_subjects):
            st.write(f"📘 **{sub['name']}** ({sub['teacher']}) — {sub['hours']} ч.")
            if st.button("❌ Удалить", key=f"sched_delete_subject_{i}_{sub['name']}"): 
                st.session_state.temp_subjects.pop(i)
                st.rerun()

    if st.button("💾 " + ("СОХРАНИТЬ ИЗМЕНЕНИЯ" if is_ed else "СОХРАНИТЬ ГРУППУ"), type="secondary", use_container_width=True):
        if not g_name or not st.session_state.temp_subjects or len(dates)!=2: 
            st.error("Заполните название, выберите период дат и добавьте хотя бы один предмет!")
        else:
            payload = {'name': g_name, 'num_courses': g_c, 'room': g_room, 'work_week': g_week, 'max_weekday_lessons': g_w_l, 'max_saturday_lessons': g_s_l, 'max_per_subject': g_m_s, 'start_date': dates[0], 'end_date': dates[1], 'subjects': list(st.session_state.temp_subjects)}
            if is_ed: st.session_state.global_groups[st.session_state.editing_idx] = payload
            else: st.session_state.global_groups.append(payload)
            st.session_state.editing_idx = None
            st.session_state.temp_subjects = []
            st.rerun()

    if st.session_state.do_scroll:
        components.html("<script>window.parent.document.getElementById('edit-section').scrollIntoView({behavior: 'smooth'});</script>", height=0)
        st.session_state.do_scroll = False